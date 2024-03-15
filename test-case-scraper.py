from polarionatofork import polarion # https://github.com/alexandertoepfer/python-polarion
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from time import sleep
from tqdm import tqdm
from itertools import islice
from io import BytesIO
from tempfile import NamedTemporaryFile
import warnings
warnings.filterwarnings("ignore", message="Unverified HTTPS request is being made") # No certificate check
import jsbeautifier
import argparse
import hashlib
import signal
import logging
import errno
import json
import time
import ast
import sys
import os
import re

# TODO: Add title check for polarion ticket with ranorex entry

class DelayedKeyboardInterrupt:

    def __enter__(self):
        self.signal_received = False
        self.old_handler = signal.signal(signal.SIGINT, self.handler)

    def handler(self, sig, frame):
        self.signal_received = (sig, frame)
        logging.debug('SIGINT received. Delaying KeyboardInterrupt.')

    def __exit__(self, type, value, traceback):
        signal.signal(signal.SIGINT, self.old_handler)
        if self.signal_received:
            self.old_handler(*self.signal_received)

parser = argparse.ArgumentParser(
    description = '''Matching between excel files from ARI and polarion tickets.''',
    epilog = "example: python scrape-tests.py \".\MyExcelFile.xlsx\" \"{\'MyPolarionUser\':\'MyAccessToken\' or \'MyTxtWithToken.txt\'}\"")
parser.add_argument('filename', nargs = 1, help = 'Path to the excel file from ARI.')
parser.add_argument('credentials', nargs = 1, help = 'Dictionary with username and access token or text file with access token. (See README.md)')
parser.add_argument('chunk', nargs = '?', help = 'Amount of tickets to be queried at once, higher values cause less responsiveness.')
args = parser.parse_args()

filename = None
username = None
token = None
chunkSize = None

def chunker(seq, size):
    return (seq[pos:pos + size] for pos in range(0, len(seq), size))

def sha256sum(data):
    m = hashlib.md5()
    m.update(data.encode())
    return m.hexdigest()

def main(filename, credentials, chunk):
    cred = json.loads(credentials.replace('\'', '"'))
    if not os.path.isfile(filename):
        raise FileNotFoundError("%s was not found or is a directory." % filename)
    filename = filename
    for k,v in cred.items():
        username = k
        if ".txt" in v:
            with open(v) as mytxt:
                token = mytxt.read()
        else:
            token = v
    chunkSize = int(chunk)

    redFill = PatternFill(start_color = 'FFFF7D7D', end_color = 'FFFF7D7D', fill_type = 'solid')
    options = jsbeautifier.default_options()
    options.indent_size = 2
    client = polarion.Polarion('REDACTED', username, password = None, token = token, verify_certificate = False)
    project = client.getProject('REDACTED')
    filenames = [filename] # Could be used for multiple files, adjust commandline args

    tickets = [] # Relevant ticket objects classified into valid/invalid testcases
    indices = [] # List of excel indices to mark
    comments = [] # List of comments for invalid tickets
    workitems = [] # Responses from polarion api wrapped into subset objects
    cells = [] # Cell values from excel sheet
    i = 0
    j = 0

    # Here be dragons, do not review this code
    parameterpattern = re.compile(
        '<(?:[^<>]*?)(?:(?:(?:class=(?:(?:\"|\')|(?:\"|\')(?:[^\"\'<>]*)\s)(?:[P|p]olarion[-_\s]?)?'
        '(?:[R|r]te[-_\s])?[t|T]est[p|P]arameter(?:(?:\"|\')|\s(?:[^\"\'<>]*)(?:\"|\')))(?:[^<>]*?)'
        '(?:data-name=(?:(?:\"|\')|(?:\"|\')(?:[^\"\'<>]*)\s)([^\"\'<>]+)(?:(?:\"|\')|(?:\"|\')'
        '(?:[^\"\'<>]*)\s)))|(?:(?:data-name=(?:(?:\"|\')|(?:\"|\')(?:[^\"\'<>]*)\s)([^\"\'<>]+)'
        '(?:(?:\"|\')|(?:\"|\')(?:[^\"\'<>]*)\s))(?:[^<>]*?)(?:class=(?:(?:\"|\')|(?:\"|\')'
        '(?:[^\"\'<>]*)\s)(?:[P|p]olarion[-_\s]?)?(?:[R|r]te[-_\s])?[t|T]est[p|P]arameter'
        '(?:(?:\"|\')|\s(?:[^\"\'<>]*)(?:\"|\')))))(?:[^<>]*?)>'
    )
    ticketidpattern = re.compile('AP[-_]\d+')
    testidpattern = re.compile('AP[-_]\d+.AP[-_]\d+')

    for filename in filenames:
        if ".xlsx" in filename:
            workbook = load_workbook(filename)
            workbook.save(filename)
            workbook.close() # Permission check on excel file
            workbook = load_workbook(filename)
            sheet = workbook.active
            headers = [c.value for c in next(sheet.iter_rows(min_row = 1, max_row = 1))]
            verifiedCol = False # Check format of excel file

            # Get all cells in Polarion ID column
            for row_cells in sheet.iter_rows(min_col = 1, max_col = 1):
                for cell in row_cells:
                    if cell.value == headers[0] == "Polarion ID": # Do nothing if header is invalid
                        verifiedCol = True
                        continue
                    if verifiedCol:
                        cells.append(cell.value)

            delete = False
            if os.path.isfile("__cache__.json"):
                with open("__cache__.json", 'rb') as f:
                    cache = json.loads(f.read())
                    if cache["filename"] == filename and cache["hash"] == sha256sum(''.join(cells)):
                        delete = True
                        workitems = cache["items"]
                        i = int(cache["index"])
                        cells = cache["cells"]

            myitem = None
            # Create chunks from cells and query Polarion
            try:
                with tqdm(total = len(cells[i::])) as pbar:
                    for group in chunker(cells[i::], chunkSize):
                        query = "id:("
                        for pId in group[:-1]:
                            query = query + pId + ' '
                        query = query + group[-1] + ')'
                        pbar.set_description("GET {q:%s...%s}" % (group[0], group[-1]))
                        items = project.searchWorkitemFullItemWithTestId(query, limit = chunkSize)
                        for item in items:
                            i = i + 1
                            workitems.append({
                                "id": item.getId(),
                                "title": item.getTitle(),
                                "type": item.getType(),
                                "author": str(item.getAuthor()),
                                "status": item.getStatus(),
                                "testid": item.getTestId(),
                                "steps": item.getTestSteps()
                            })
                            tqdm.write('#' + str(i - 1).zfill(3) + ' {' + 'i:' + item.getId() + ', t:' + item.getTitle() + '}')
                        pbar.update(len(items))
            except KeyboardInterrupt:
                with DelayedKeyboardInterrupt():
                    with open('__cache__.json', 'w') as out:
                        res = {
                            "filename": filename,
                            "hash": sha256sum(''.join(cells)),
                            "index": i,
                            "cells": cells,
                            "items": workitems
                        }
                        res = jsbeautifier.beautify(json.dumps(res), options)
                        print(res, file = out)
                        quit()

            i = 0
            verifiedCol = False
            # Verify customFields.testCaseID from workitems and parse test parameters
            with tqdm(total = len(sheet['A']) - 1) as pbar:
                for row_cells in sheet.iter_rows(min_col = 1, max_col = 1):
                    for cell in row_cells:
                        pbar.update(1)
                        if cell.value == headers[0] == "Polarion ID":
                            verifiedCol = True
                            continue
                        if verifiedCol:
                            i = i + 1
                            workitem, = [x for x in workitems if x["id"] == cell.value]
                            pbar.set_description("READ {w:%s}" % workitem["id"])
                            time.sleep(0.025) # For tqdm buffer to catch up, to not overwrite printed characters
                            # Simulate one id which doesn't match
                            #if workitem.getId() == 'AP-101163':
                            #    workitem.setTestId("AP-101164.AP-101164 Something")
                            #    workitem.setType("testcase")
                            if workitem["type"] != "testcase" or workitem["testid"] is None:
                                cell.fill = redFill
                                # +1 Offset from excel header, keep this for future adjustments in case format changes
                                indices.append(i - 1 + 1)
                                comments.append("Ticket is not a testcase." if workitem["type"] != "testcase" else "Ticket does not have a testCaseID.")
                                tickets.append({"title": workitem["id"] + ' ' + workitem["title"], "type": workitem["type"], "author": workitem["author"], "status": workitem["status"], "comment": comments[-1], "id": "", "params": []})
                                tqdm.write('#' + str(i - 1).zfill(3) + ' {' + 'w:' + workitem["id"] + ' ' + workitem["title"] + ', i:None ' + ', p:None' + '}')
                                continue

                            # Set customFields.testCaseID
                            tid = "" if workitem["testid"] is None else workitem["testid"]

                            match = False
                            # Check the first two AP numbers inside customFields.testCaseID to match the ticket
                            for y in islice(re.finditer(ticketidpattern, workitem["testid"]), 2):
                                if y is None:
                                    continue
                                else:
                                    if y.group() == workitem["id"]:
                                        match = True

                            # Not sure if this extra check is really needed
                            if workitem["id"] + "." + workitem["id"] + " - " + workitem["title"] == workitem["testid"] or workitem["id"] + "." + workitem["id"] == workitem["testid"]:
                                match = True
                            else:
                                match = False

                            if '"' in workitem["title"]:
                                match = False

                            # Treat as invalid ticket in case AP numbers are missing or invalid, parameters don't matter
                            if not match:
                                cell.fill = redFill
                                indices.append(i - 1 + 1)
                                comments.append("Ticket contains invalid testCaseID")
                                tickets.append({"title": workitem["id"] + ' ' + workitem["title"], "type": workitem["type"], "author": workitem["author"], "status": workitem["status"], "comment": comments[-1], "id": tid, "params": []})
                                tqdm.write('#' + str(i - 1).zfill(3) + ' {' + 'w:' + workitem["id"] + ' ' + workitem["title"] + ', i:' + tid + ', p:None' + '}')
                                continue

                            params = {}
                            # Parse HTML to check for test parameters
                            for step in workitem["steps"]:
                                for key in step.keys():
                                    if step[key] is None:
                                        continue
                                    for y in re.findall(parameterpattern, step[key]):
                                        if y is None:
                                            continue
                                        else:
                                            for x in y:
                                                if x:
                                                    params[x] = ""

                            # Treat as invalid ticket in case test parameters were found
                            if any(params):
                                cell.fill = redFill
                                indices.append(i - 1 + 1)
                                comments.append("Ticket contains test parameters")
                                tickets.append({"title": workitem["id"] + ' ' + workitem["title"], "type": workitem["type"], "author": workitem["author"], "status": workitem["status"], "comment": comments[-1], "id": tid, "params": list(params.keys())})
                                tqdm.write('#' + str(i - 1).zfill(3) + ' {' + 'w:' + workitem["id"] + ' ' + workitem["title"] + ', i:' + tid + ', p:' + ''.join(list(params.keys())) + '}')
                                continue

                            # Else a correct ticket entry with customFields.testCaseID will be created with a presumably empty test parameter list
                            tickets.append({"title": workitem["id"] + ' ' + workitem["title"], "type": workitem["type"], "author": workitem["author"], "status": workitem["status"], "id": tid, "params": list(params.keys())})
                            tqdm.write('#' + str(i - 1).zfill(3) + ' {' + 'w:' + workitem["id"] + ' ' + workitem["title"] + ', i:' + tid + ', p:' + ''.join(list(params.keys())) + '}')
            i = 0
            verifiedCol = False
            # Add information about invalid tickets to excel as notes
            for row_cells in sheet.iter_rows(min_col = 2, max_col = 2):
                for cell in row_cells:
                    j = j + 1
                    if cell.value == headers[1] == "Test case name":
                        verifiedCol = True
                        continue
                    if verifiedCol:
                        if (j - 1) in indices:
                            i = i + 1
                            cell.fill = redFill
                            cell.comment = Comment(comments[i - 1], "Test Automation")
            workbook.save(filename)
            workbook.close()

    res = []
    # Re-sort list to write to json file
    sorted(tickets, key = lambda k: (k["id"]))
    for dict in tickets:
        if dict["id"] and not dict["params"]:
            res.append({k:dict[k] for k in ('title', 'id')})
    for dict in tickets:
        if not dict["id"] or dict["params"]:
            res.append(dict)

    # Write json report from collected data
    #with open('tests.json', 'w') as out:
    #    res = jsbeautifier.beautify(json.dumps(res), options)
    #    print(res, file = out)

    if delete:
        os.remove("__cache__.json")

if __name__ == '__main__':
    if len(sys.argv) > 3:
        main(str(sys.argv[1]), str(sys.argv[2]), str(sys.argv[3]))
    else:
        main(str(sys.argv[1]), str(sys.argv[2]), str(15))

# TODO: Implement automatic error correction techniques such as confidence score near exact queries for title/number/type mismatches.
# Write actions are dangerous on live systems, would be better with bot accounts on dummy tickets cut off from live systems (to not destroy anything on live)
